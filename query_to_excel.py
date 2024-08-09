import pandas as pd
import time
from openpyxl import load_workbook
from openpyxl.worksheet.table import Table, TableStyleInfo
from openpyxl.styles import Alignment,Font,PatternFill, Border, Side

def query_to_excel(conn, query, excel_file, sheet_name,title):
    try:
        print(f"Creando Archivo en la hoja {sheet_name}")
        df = pd.read_sql_query(query, conn)

        columns_to_replace = ['Total $', 'Total Bs']
        # df = replace_dot_with_comma(df, columns_to_replace)

        with pd.ExcelWriter(excel_file, engine='openpyxl', mode='w') as writer:
            df.to_excel(writer, index=False, sheet_name=sheet_name)


        libro = load_workbook(excel_file)
        hoja = libro[sheet_name]

        hoja.insert_rows(1)

        hoja.cell(row=1, column=1, value=title)

        hoja.merge_cells(f'A1:{chr(64 + len(df.columns))}1')

        title_cell = hoja['A1']
        title_cell.font = Font(bold=True, size=12, color="000000")  # Fuente en negrita, tamaño 16, color blanco
        title_cell.fill = PatternFill(start_color="8DB4E2", end_color="8DB4E2", fill_type="solid")  # Fondo azul marino
        title_cell.alignment = Alignment(horizontal='center')
        border = Side(border_style="thin", color="000000")
        title_cell.border = Border(left=border, right=border, top=border, bottom=border)

        tab_range = f"A2:{chr(64 + len(df.columns))}{len(df) + 2}"

        tab = Table(displayName="Tabla1", ref=tab_range)

        style = TableStyleInfo(
            name="TableStyleMedium9", 
            showFirstColumn=False,
            showLastColumn=False, 
            showRowStripes=True, 
            showColumnStripes=True
        )
        tab.tableStyleInfo = style

        hoja.add_table(tab)

        for col in columns_to_replace:
            if col in df.columns:
                col_idx = df.columns.get_loc(col) + 1  # Obtener el índice de la columna
                for cell in hoja.iter_cols(min_col=col_idx, max_col=col_idx, min_row=2, max_row=len(df)+1):
                    for c in cell:
                        c.number_format = '#,##0.00'  # Formato con separador de miles

        if 'Monto Total' in df.columns:
            if 'Numero de Documento' in df.columns:
                for cell in hoja['C'][2:]:
                    cell.alignment = Alignment(horizontal='center')
                for cell in hoja['D'][2:]:
                    cell.alignment = Alignment(horizontal='center')
                for cell in hoja['E'][2:]:
                    cell.alignment = Alignment(horizontal='center')
            else:
                for cell in hoja['C'][2:]:  
                    cell.alignment = Alignment(horizontal='center')
                


        for column in hoja.columns:
            max_length = 0
            column_letter = column[1].column_letter
            for cell in column:
                try:
                    if cell.value:
                        cell_length = len(str(cell.value))
                        if cell_length > max_length:
                            max_length = cell_length
                except:
                    pass

            hoja.column_dimensions[column_letter].width = max_length + 2

        libro.save(excel_file)
        print(f"Resultados guardados en {excel_file}")
        time.sleep(2)
    except Exception as e:
        print(f"Error al ejecutar la consulta y guardar en Excel: {e}")
        time.sleep(2)
